import argparse
import os
import openpyxl
from datetime import datetime, time, timedelta
from typing import List, Tuple, Optional, Any
from openpyxl import workbook, worksheet

ROW_FIRST: int = 7
ROW_LAST: int = 38
# ['A': 日, [(朝) 'D': 時刻, 'E': 最高血圧, 'G': 最低血圧, 'I': 脈拍]
#         , [(夜) 'K': 時刻, 'L': 最高血圧, 'N': 最低血圧, 'P': 脈拍]
# 'S': 歩数  ]
COL_NAMES: List[str] = ['A', 'D', 'E', 'G', 'I', 'K', 'L', 'N', 'P', 'S']
DATE_COL_NAME: str = "A"
TIME_COL_NAMES: Tuple[str] = ('D', 'K')
STEP_COL_NAME: str = "S"
# PID,測定日付,(朝)時刻,(朝)最高血圧,(朝)最低血圧,(朝)脈拍,(夜)最高血圧,(夜)最低血圧,(夜)脈拍
FMT_LINE_BLOOD_PRESSURE: str = '{0},"{1}",{2},{3},{4},{5},{6},{7},{8},{9}\n'
# PID,測定日付,歩数
FMT_LINE_STEP_COUNT: str = '{0},"{1}",{2}\n'
# OUTPUT CSV HEADER
_HDR_MORNING: str = '"morning_measurement_time","morning_max","morning_min","morning_pulse_rate",'
_HDR_NIGHT: str = '"evening_measurement_time","evening_max","evening_min","evening_pulse_rate"\n'
HEADER_BLOOD_PRESSURE: str = '"pid","measurement_day",' + _HDR_MORNING + _HDR_NIGHT
# 歩数CSV HEADER
HEADER_WALKING_COUNT: str = '"pid","measurement_day","counts"\n'
# OUTPUT CSV FILE
OUTPUT_BLOOD_PRESSURE: str = "blood_pressure-{sheet_name}.csv"
OUTPUT_STEP_COUNT: str = "walking_count-{sheet_name}.csv"


def _save_csv(file_path: str, data: List[str]):
    with open(file_path, mode='w') as fp:
        line_cnt = 0
        for line in data:
            fp.write(line)
            line_cnt += 1
    return line_cnt


if __name__ == '__main__':
    parser: argparse.ArgumentParser = argparse.ArgumentParser()
    # Excel bookファイルパス: 必須
    # (例) ~/Documents/Excel/血圧測定表(2022).xlsx"
    parser.add_argument("--file-path", type=str, required=True,
                        help="Blood pressure and step Excel book path.")
    # シート名: 必須
    parser.add_argument("--sheet-name", type=str, required=True,
                        help="Output sheet name.")
    args: argparse.Namespace = parser.parse_args()

    # CSVフアイルパス: 先頭 "~" はユーザホームディレクトリ展開する
    excel_file: str = os.path.join(os.path.expanduser(args.file_path))
    if not os.path.exists(excel_file):
        print(f"Excel book [{excel_file}] is not found!")
        exit(1)

    excel_file = os.path.expanduser(excel_file)
    # data_only=True: 式を無視して値のみ取得
    wkbk: workbook  = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws: worksheet = wkbk[args.sheet_name]
    # 年月
    yearMonth: datetime = ws['A3'].value
    # pid
    pid: str = ws['E3'].value
    # 血圧測定CSV
    csv_blood_list: List[str] = [HEADER_BLOOD_PRESSURE]
    # 歩数測定CSV
    csv_walking_list: List[str] = [HEADER_WALKING_COUNT]
    # 行データ無しフラグ
    is_row_empty = False
    for row_num in range(ROW_FIRST, ROW_LAST):
        cols: List[str] = [col_name + str(row_num) for col_name in COL_NAMES]
        blood_list: List[str] = [pid]
        step_list: List[str] = [pid]
        for col in cols:
            col_value: Optional[Any]  = ws[col].value
            if col.startswith(DATE_COL_NAME):
                if not isinstance(col_value, int):
                    # 日セルが空なら終了
                    is_row_empty = True
                    break

                # 測定日付作成
                date_num: int = int(ws[col].value) - 1
                date_col: datetime = yearMonth + timedelta(days=date_num)
                str_date_value: str = date_col.strftime("%Y-%m-%d")
                blood_list.append(str_date_value)
                step_list.append(str_date_value)
            elif col.startswith(TIME_COL_NAMES):
                # 時刻
                str_time_value: str
                if isinstance(col_value, time):
                    time_col: datetime = ws[col].value
                    str_time_value = time_col.strftime("%H:%M")
                    str_time_value = '"' + str_time_value + '"'
                else:
                    # 欠測データ時刻: ":"
                    str_time_value = ""
                blood_list.append(str_time_value)
            elif col.startswith(STEP_COL_NAME):
                # 歩数
                step_list.append(ws[col].value)
            else:
                # 数値データ: 最高血圧, 最低血圧, 脈拍
                any_value: Any = ws[col].value
                if any_value != "-":
                    value: int = any_value
                else:
                    # 欠測値: "-"
                    value = ""
                blood_list.append(value)
        if not is_row_empty:
            print(f"{row_num} {blood_list}")
            line_blood: str = FMT_LINE_BLOOD_PRESSURE.format(*blood_list)
            csv_blood_list.append(line_blood)
            line_step: str = FMT_LINE_STEP_COUNT.format(*step_list)
            csv_walking_list.append(line_step)

    # 保存パス
    save_path: str = os.path.dirname(excel_file)
    # 血圧測定CSVファイル
    blood_name: str = OUTPUT_BLOOD_PRESSURE.format(sheet_name=args.sheet_name)
    blood_path: str = os.path.join(save_path, "csv", blood_name)
    _save_csv(blood_path, csv_blood_list)
    print(f"Saved: {blood_path}")
    # 歩数測定CSVファイル
    step_name: str = OUTPUT_STEP_COUNT.format(sheet_name=args.sheet_name)
    step_path: str = os.path.join(save_path, "csv", step_name)
    _save_csv(step_path, csv_walking_list)
    print(f"Saved: {step_path}")
