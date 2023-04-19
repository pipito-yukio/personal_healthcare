import argparse
import os
import openpyxl
from datetime import datetime, time, timedelta
from typing import List, Tuple, Optional, Any
from openpyxl import workbook, worksheet

"""
健康管理(睡眠、夜間頻尿要因、天候)Excelシートから下記テーブル用CSVをエクスポートする
1.健康管理データベース: healthcare_db
 1-1.睡眠管理テーブル用: bodyhealth.sleep_management -> sleep_management-YYYYmm.csv
 1-2.夜間頻尿要因: bodyhealth.nocturia_factors -> nocturia_factors-YYYYmm.csv
 1-3.体温測定: bodyhealth.body_temperature -> body_temperature-YYYYmm.csv
     これは主キーのみ残りのカラムはnull
2.気象センサーデータベース: sensors_pgdb
 2-1.天候テーブル 
"""


ROW_FIRST: int = 7
ROW_LAST: int = 38
# 共通
# ['A: 日]
# 睡眠管理
# ['C': 起床時間, 'D': 睡眠スコア, 'E': 睡眠時間, 'F': 深い睡眠]
# 夜間頻尿要因
# ['G':トイレ回数,'H': 珈琲,'I':お茶,'J':酒,'K':栄養ド,'L':スポド,'M':その他,'N':服薬 'O':入浴,'P':健康状態]
# 天候
# ['Q': 天気 ]
COL_NAMES: List[str] = ['A','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
DATE_COL_NAME: str = "A"
# 睡眠管理列名リスト
SLEEP_MAN_NAMES: Tuple[str] = ('C','D','E','F')
TIME_COL_NAMES: Tuple[str] = ('C','E','F')
# 夜間頻尿要因列名リスト
NOCTURIA_FACT_NAMES: Tuple[str] = ('G','H','I','J','K','L','M','N','O','P')
#  数値(この項目のみ)
VISITS_COL_NAME: str = "G"
HEALTH_CONDITION_COL_NAME: str = "P"
# 天候
WEATHER_COL_NAME: str = "Q"
# PID,測定日付,起床時間,睡眠スコア,睡眠時間,深い睡眠
FMT_LINE_SLEEP_MAN: str = '{0},"{1}",{2},{3},{4},{5}\n'
# true/false 項目 (3-10)は文字列フォーマットでクォートする
# 健康状態[11]: ブランクの場合なにも出力しない(null)
#  値が有る場合はプログラムでダブルクォートでくくる
FMT_LINE_NOCTURIA_FACT: str = '{0},"{1}",{2},"{3}","{4}","{5}","{6}","{7}","{8}","{9}","{10}",{11}\n'
# 体温測定: 主キー以外はnull
FMT_LINE_BODY_TEMPER: str = '{0},"{1}",,\n'
# 測定日付,天候はNOT NULL なので必ず入力される
FMT_LINE_WEATHER: str = '"{0}","{1}"\n'
# OUTPUT CSV HEADER
# 睡眠管理
HEADER_SLEEP_MAN: str = '"pid","measurement_day","wakeup_time","sleep_score","sleeping_time","deep_sleeping_time"\n'
# 夜間頻尿要因
_FACT1: str = '"has_coffee","has_tea","has_alcohol","has_nutrition_drink","has_sports_drink","has_diuretic",'
_FACT2: str = '"take_medicine","take_bathing","condition_memo"\n'
HEADER_NOCTURIA_FACT: str = '"pid","measurement_day","midnight_toilet_visits"' + _FACT1 + _FACT2
# 体温測定
HEADER_BODY_TEMPER: str = '"pid","measurement_day","measurement_time","temperature"\n'
# 天候CSV HEADER
HEADER_WEATHER: str = '"measurement_day","condition"\n'
# OUTPUT CSV FILE
OUTPUT_SLEEP_MAN: str = "sleep_management-{sheet_name}.csv"
OUTPUT_NOCTURIA_FACT: str = "nocturia_factors-{sheet_name}.csv"
OUTPUT_BODY_TEMPER: str = "body_temper-{sheet_name}.csv"
OUTPUT_WEATHER: str = "weather_condition-{sheet_name}.csv"


def _save_csv(file_path: str, data: List[str]):
    with open(file_path, mode='w') as fp:
        line_cnt = 0
        for line in data:
            fp.write(line)
            line_cnt += 1
    return line_cnt


def stringfy(val: str) -> str:
    return '\"' + val + '\"'


if __name__ == '__main__':
    parser: argparse.ArgumentParser = argparse.ArgumentParser()
    # Excel bookファイルパス: 必須
    # (例) ~/Documents/Excel/睡眠管理_2023.xlsx"
    parser.add_argument("--file-path", type=str, required=True,
                        help="睡眠管理 Excel book path.")
    # シート名: 必須
    parser.add_argument("--sheet-name", type=str, required=True,
                        help="Output sheet name.")
    args: argparse.Namespace = parser.parse_args()

    # CSVフアイルパス: 先頭 "~" はユーザホームディレクトリ展開する
    excel_file: str = os.path.join(os.path.expanduser(args.file_path))
    if not os.path.exists(excel_file):
        print(f"Excel book [{excel_file}] is not found!")
        exit(1)

    excel_file = os.path.expanduser(excel_file)
    # data_only=True: 式を無視して値のみ取得
    wkbk: workbook  = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws: worksheet = wkbk[args.sheet_name]
    # 年月
    yearMonth: datetime = ws['A3'].value
    # pid
    pid: str = ws['D3'].value
    print(f"year-month:{yearMonth},pid:{pid}")
    # 睡眠管理CSV
    csv_sleep_man: List[str] = [HEADER_SLEEP_MAN]
    # 夜間頻尿要因CSV
    csv_nocturia_fact:  List[str] = [HEADER_NOCTURIA_FACT]
    # 体温CSV
    csv_body_temper: List[str] = [HEADER_BODY_TEMPER]
    # 天候CSV
    csv_weather_list: List[str] = [HEADER_WEATHER]
    # 行データ無しフラグ
    is_row_empty = False
    for row_num in range(ROW_FIRST, ROW_LAST):
        cols: List[str] = [col_name + str(row_num) for col_name in COL_NAMES]
        sleep_man_list: List[str] = [pid]
        nocturia_fact_list: List[str] = [pid]
        body_temper_list: List[str] = [pid]
        # 天候は日付のみ
        weather_list: List[str] = []
        for col in cols:
            col_value: Optional[Any]  = ws[col].value
            if col.startswith(DATE_COL_NAME):
                if not isinstance(col_value, int):
                    # 日セルが空なら終了
                    is_row_empty = True
                    break

                # 測定日付作成
                date_num: int = int(ws[col].value) - 1
                date_col: datetime = yearMonth + timedelta(days=date_num)
                str_date_value: str = date_col.strftime("%Y-%m-%d")
                sleep_man_list.append(str_date_value)
                nocturia_fact_list.append(str_date_value)
                body_temper_list.append(str_date_value)
                weather_list.append(str_date_value)
            elif col.startswith(SLEEP_MAN_NAMES):
                # 睡眠管理: 起床時刻,睡眠スコア,睡眠時間,深い睡眠
                #   時刻
                if col.startswith(TIME_COL_NAMES):
                    str_time_value: str
                    if isinstance(col_value, time):
                        time_col: datetime = ws[col].value
                        str_time_value = time_col.strftime("%H:%M")
                        str_time_value = '"' + str_time_value + '"'
                    else:
                        # 欠測データ時刻: ":"
                        str_time_value = ""
                    sleep_man_list.append(str_time_value)
                else:
                    # 睡眠スコア: 空の場合が有る
                    any_value: Any = ws[col].value
                    if any_value != "-":
                        value: int = any_value
                    else:
                        # 欠測値: "-"
                        value: str = ""
                    sleep_man_list.append(value)
            elif col.startswith(NOCTURIA_FACT_NAMES):
                # 夜間頻尿要因
                if col.startswith(VISITS_COL_NAME):
                    # トイレ回数: 整数 (これのみ)
                    nocturia_fact_list.append(ws[col].value)
                elif col.startswith(HEALTH_CONDITION_COL_NAME):
                    # 健康状態
                    any_value: Any = ws[col].value
                    if any_value is not None:
                        # 入力されている場合はプログラムでダブルクォートでくくる
                        value: str = stringfy(any_value)
                        nocturia_fact_list.append(value)
                    else:
                        # 未入力の場合はブランクとする
                        nocturia_fact_list.append("")
                else:
                    # 1(true), ''(false)
                    any_value: Any = ws[col].value
                    if any_value is not None:
                        value: str = "t"
                    else:
                        # 未入力は全て: "false"
                        value: str = "f"
                    nocturia_fact_list.append(value)
            elif col.startswith(WEATHER_COL_NAME):
                # 天候: 必須項目なので必ずある
                weather_list.append(ws[col].value)
        if not is_row_empty:
            print(f"{row_num} {sleep_man_list}")
            print(f"{row_num} {nocturia_fact_list}")
            print(f"{row_num} {weather_list}")
            line_sleep_man: str = FMT_LINE_SLEEP_MAN.format(*sleep_man_list)
            csv_sleep_man.append(line_sleep_man)
            line_nocturia_fact: str = FMT_LINE_NOCTURIA_FACT.format(*nocturia_fact_list)
            csv_nocturia_fact.append(line_nocturia_fact)
            line_body_temper: str = FMT_LINE_BODY_TEMPER.format(*body_temper_list)
            csv_body_temper.append(line_body_temper)
            line_weather: str = FMT_LINE_WEATHER.format(*weather_list)
            csv_weather_list.append(line_weather)

    # Excelファイルのパス
    save_base_path: str = os.path.dirname(excel_file)
    # CSVの保存パス: Excelファイルパスの"csv"サブディレクトリ
    save_csv_path: str = os.path.join(save_base_path, "csv")
    if not os.path.exists(save_csv_path):
        os.makedirs(save_csv_path)

    # 睡眠管理CSVファイル
    sleep_man_name: str = OUTPUT_SLEEP_MAN.format(sheet_name=args.sheet_name)
    sleep_man_path: str = os.path.join(save_csv_path, sleep_man_name)
    _save_csv(sleep_man_path, csv_sleep_man)
    print(f"Saved: {sleep_man_path}")
    # 夜間頻尿要因CSVファイル
    nocturia_fact_name: str = OUTPUT_NOCTURIA_FACT.format(sheet_name=args.sheet_name)
    nocturia_fact_path: str = os.path.join(save_csv_path, nocturia_fact_name)
    _save_csv(nocturia_fact_path, csv_nocturia_fact)
    print(f"Saved: {nocturia_fact_path}")
    # 体温測定CSVファイル
    body_temper_name: str = OUTPUT_BODY_TEMPER.format(sheet_name=args.sheet_name)
    body_temper_path: str = os.path.join(save_csv_path, body_temper_name)
    _save_csv(body_temper_path, csv_body_temper)
    print(f"Saved: {body_temper_path}")
    # 天候CSVファイル
    weather_name: str = OUTPUT_WEATHER.format(sheet_name=args.sheet_name)
    weather_path: str = os.path.join(save_csv_path, weather_name)
    _save_csv(weather_path, csv_weather_list)
    print(f"Saved: {weather_path}")
